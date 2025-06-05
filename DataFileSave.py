import openpyxl
from openpyxl import Workbook

class DataFileSave:
    def __init__(self):

        self.data_io_for_modbus = []
        self.data_io_var = []
        self.data_io = []
        self.data_io_type = []
        self.data_io_product = []
        self.data_io_min = []
        self.data_io_max = []

        self.converter_count_input = 0
        self.converter_count_output = 0
        self.converter_modbus_use = False
        self.converter_type_current = 1
        self.converter_data_for_modbus = ()

        self.heat1_type = 0
        self.heat2_use = False
        self.heat2_type = 0
        self.heat_data_for_modbus = ()

        self.recup_use = False
        self.recup_type = 0
        self.recup_data_for_modbus = ()

        self.dx_use = False
        self.dx_type = 0
        self.dx_data_for_modbus = ()

        self.humidifier_use = False
        self.humidifier_data_for_modbus = ()

        self.mix_camera_use = False
        self.mix_camera_data_for_modbus = ()

        self.modbus_name_using_sensors = []
        self.modbus_name_type_sensors = []
        self.modbus_var_sensors = []
        self.modbus_data_for_sensors = []

    def setIOData(self, io, io_var, io_type, io_product, io_min, io_max, io_modbus=0):
        self.data_io_for_modbus = io_modbus
        self.data_io_var = io_var
        self.data_io = io
        self.data_io_type = io_type
        self.data_io_product = io_product
        self.data_io_min = io_min
        self.data_io_max = io_max

    def setConverterData(self, count_input, count_output, modbus_use, type_current, data_modbus=0):
        self.converter_count_input = count_input
        self.converter_count_output = count_output
        self.converter_modbus_use = bool(modbus_use)
        self.converter_type_current = type_current
        self.converter_data_for_modbus = data_modbus

    def setHeatData(self, heat1_type, heat2_use, heat2_type, data_for_modbus=0):
        self.heat1_type = heat1_type
        self.heat2_use = bool(heat2_use)
        self.heat2_type = heat2_type
        self.heat_data_for_modbus = data_for_modbus

    def setRecupData(self, recup_use, recup_type, recup_data=0):
        self.recup_use = bool(recup_use)
        self.recup_type = recup_type
        self.recup_data_for_modbus = recup_data

    def setDxData(self, dx_use, dx_type, dx_data=0):
        self.dx_use = bool(dx_use)
        self.dx_type = dx_type
        self.dx_data_for_modbus = dx_data

    def setHumidifierData(self, hum_use, hum_data=0):
        self.humidifier_use = bool(hum_use)
        self.humidifier_data_for_modbus = hum_data

    def setMixCameraData(self, mix_use, mix_data=0):
        self.mix_camera_use = bool(mix_use)
        self.mix_camera_data_for_modbus = mix_data

    def setModbusSensorsData(self, name_sensor, type_sensor, var_sensors, sourse, id_list=0, data_for_modbus=0):
        if (sourse == "eplan"):
            self.modbus_name_using_sensors = name_sensor
            self.modbus_name_type_sensors = type_sensor
            self.modbus_var_sensors = var_sensors
            self.modbus_data_for_sensors = data_for_modbus
        else:
            self.id_sensors_list = id_list
            self.modbus_name_type_sensors = name_sensor
            self.modbus_var_sensors = var_sensors

    def createFile(self, file_path, sourse):
        self.wb = openpyxl.load_workbook(file_path)
        ws = self.wb.active
        ws.title = "Данные для плк"


        self.createDataIO(ws, sourse)
        self.createDataConverter(ws, sourse)
        self.createDataHeat(ws, sourse)
        self.createDataRecup(ws, sourse)
        self.createDataDx(ws, sourse)
        self.createDataHumidifier(ws, sourse)
        self.createDataMix(ws, sourse)
        self.createDataSensor(ws, sourse)

        self.wb.save(file_path)


    def createDataIO(self, ws, sourse):
        ws[f"A{1}"] = "Номер контакта"
        ws[f"B{1}"] = "Переменная"
        ws[f"C{1}"] = "Тип"
        ws[f"D{1}"] = "Метод"
        ws[f"E{1}"] = "Min"
        ws[f"F{1}"] = "Max"
        if (sourse == 'eplan'):
            for i in range(0, len(self.data_io)):
                ws[f"A{i + 2}"] = self.data_io[i]
                ws[f"B{i + 2}"] = self.data_io_var[i][0]
                str = self.tuples_to_string(self.data_io_for_modbus[i])
                ws[f"C{i + 2}"] = str
        if (sourse == 'plc'):
            for i in range(0, len(self.data_io)):
                ws[f"A{i + 2}"] = self.data_io[i]
                ws[f"B{i + 2}"] = self.data_io_var[i]
                if (i < 18):
                    ws[f"D{i + 2}"] = self.data_io_product[i]
                if (i < 25):
                    ws[f"C{i + 2}"] = self.data_io_type[i]
                if (i < 6):
                    ws[f"E{i + 2}"] = self.data_io_min[i]
                    ws[f"F{i + 2}"] = self.data_io_max[i]

    def createDataConverter(self, ws, sourse):
        ws['I1'] = "Количество ПЧ приток:"
        ws['J1'] = self.converter_count_input

        ws['I2'] = "Количество ПЧ вытяжка:"
        ws['J2'] = self.converter_count_output

        ws['I3'] = "Modbus используется:"
        ws['J3'] = self.converter_modbus_use

        ws['I4'] = "Тип ПЧ:"
        ws['J4'] = self.converter_type_current

        if (sourse == 'eplan'):
            ws['I5'] = "Данные для modbus:"
            ws['J5'] = self.tuples_to_string(self.converter_data_for_modbus)

    def createDataHeat(self, ws, sourse):
        ws['L1'] = "Тип 1 нагревателя:"
        ws['M1'] = self.heat1_type

        ws['L2'] = "Нагреватель 2 используется:"
        ws['M2'] = self.heat2_use

        ws['L3'] = "Тип 2 нагревателя:"
        ws['M3'] = self.heat2_type

        if (sourse == 'eplan'):
            ws['L4'] = "Данные для modbus:"
            ws['M4'] = self.tuples_to_string(self.heat_data_for_modbus)

    def createDataRecup(self, ws, sourse):
        ws['i7'] = "Рекуператор используется:"
        ws['J7'] = self.recup_use

        ws['I8'] = "Тип рекуператора:"
        ws['J8'] = self.recup_type

        if (sourse == 'eplan'):
            ws['I9'] = "Данные для modbus:"
            ws['FJ9'] = self.tuples_to_string(self.recup_data_for_modbus)

    def createDataDx(self, ws, sourse):
        ws['I11'] = "Охладитель используется:"
        ws['J11'] = self.dx_use

        ws['I12'] = "Тип охладителя:"
        ws['J12'] = self.dx_type

        if (sourse == 'eplan'):
            ws['I13'] = "Данные для modbus:"
            ws['J13'] = self.tuples_to_string(self.dx_data_for_modbus)

    def createDataHumidifier(self, ws, sourse):
        ws['I15'] = "Увлажнитель используется:"
        ws['J15'] = self.humidifier_use

        if (sourse == 'eplan'):
            ws['I16'] = "Данные для modbus:"
            ws['J16'] = self.tuples_to_string(self.humidifier_data_for_modbus)


    def createDataMix(self, ws, sourse):
        ws['I18'] = "Камера смешения используется:"
        ws['J18'] = self.mix_camera_use

        if (sourse == 'eplan'):
            ws['I19'] = "Данные для modbus:"
            ws['J19'] = self.tuples_to_string(self.mix_camera_data_for_modbus)

    def createDataSensor(self, ws, sourse):
        num = 0
        ws['I21'] = "Modbus датчики:"
        if (sourse == 'eplan'):
            for i in range(0, len(self.modbus_name_using_sensors)):
                ws[f"I{22 + i}"] = self.modbus_name_using_sensors[i]
                ws[f"J{22 + i}"] = self.modbus_name_type_sensors[i]
                ws[f"K{22 + i}"] = self.modbus_var_sensors[i]
                num = i + 1
                ws[f"L{22 + num}"] = self.tuples_to_string(self.modbus_data_for_sensors)

        if (sourse == 'plc'):
            ws['I22'] = "id modbus датчика:"
            ws['J22'] = "Тип modbus датчика:"
            ws['K22'] = "Переменная датчика:"
            for i in range(0, len(self.id_sensors_list)):
                ws[f"I{23 + i}"] = self.id_sensors_list[i]
                ws[f"J{23 + i}"] = self.modbus_name_type_sensors[i]
                ws[f"K{23 + i}"] = self.modbus_var_sensors[i]

    def tuples_to_string(self, tuples_list, separator=', '):
        # Преобразуем каждый кортеж в строку и объединяем их с помощью разделителя
        return separator.join(str(tup) for tup in tuples_list)
