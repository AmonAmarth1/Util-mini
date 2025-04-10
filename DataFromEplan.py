
import openpyxl

class DataFromEplan:

    def __init__(self):
        self.io = {0: ""}
        self.name_var_scheme = {0: ""}
        self.product_number_io = {0: ""}
        self.file_length_scheme_plc = 0

        self.io.clear()
        self.name_var_scheme.clear()

        self.name_var_specification = {0: ""}
        self.product_number = {0: ""}
        self.file_length_specification = 0

        self.name_var_specification.clear()
        self.product_number.clear()

    def clear(self):
        self.io.clear()
        self.name_var_scheme.clear()
        self.product_number_io.clear()
        self.file_length_scheme_plc = 0
        self.name_var_specification.clear()
        self.product_number.clear()
    def makeData(self):
        self.print_specification()

        self.setProduct_Number_IO()

        print(self.product_number_io)
        self.print_scheme_plc()

        print(self.getNameVarScheme())
        print(self.getNameVarSpecification())

    def setProduct_Number_IO(self):
        for i in range(0, self.file_length_scheme_plc):
            if f"{self.name_var_scheme[i]}" in self.name_var_specification:
                num = self.name_var_specification[f"{self.name_var_scheme[i]}"]
                self.product_number_io[i] = self.product_number[num]
            else: self.product_number_io[i] = None

    def readExel_scheme_plc(self, file_path_scheme_plc):

        wb = openpyxl.load_workbook(file_path_scheme_plc)
        sheet_name = wb.sheetnames
        sheet = wb[sheet_name[0]]

        for i in range(1, sheet.max_row):
            io = sheet[f"A{i}"]
            name_var = sheet[f"B{i}"]
            if (name_var.value != None):
                self.name_var_scheme[self.file_length_scheme_plc] = name_var.value
                self.io[self.file_length_scheme_plc] = io.value
                self.file_length_scheme_plc = self.file_length_scheme_plc + 1


    def readExel_specification(self, file_path_specification):
        wb = openpyxl.load_workbook(file_path_specification)
        sheet_name = wb.sheetnames
        sheet = wb[sheet_name[0]]
        for i in range(2, sheet.max_row + 1):
            name_var = sheet[f"C{i}"]
            product_number = sheet[f"E{i}"]
            if (name_var.value != None):
                if (name_var.value.find(" ") == -1):
                    self.name_var_specification[name_var.value] = self.file_length_specification
                    self.product_number[self.file_length_specification] = product_number.value
                    self.file_length_specification = self.file_length_specification + 1
                else:
                    words = name_var.value.split()
                    for j in range(0, len(words)):
                        self.name_var_specification[words[j]] = self.file_length_specification
                        self.product_number[self.file_length_specification] = product_number.value
                        self.file_length_specification = self.file_length_specification + 1
    def print_scheme_plc(self):
        print(self.io)
        print(self.name_var_scheme)
        print(f"length file scheme plc: {self.file_length_scheme_plc}")

    def print_specification(self):
        print(self.name_var_specification)
        print(self.product_number)
        print(f"length file specification: {self.file_length_specification}")

    def getVar(self, i):
        return self.name_var_scheme[i]

    def getIO(self, i):
        return self.io[i]

    def getProduct_number_IO(self, i):
        return  self.product_number_io[i]

    def getFileLengthSchemePlc(self):
        return self.file_length_scheme_plc

    def getFileLengthSpecification(self):
        return self.file_length_specification

    def getNameVarScheme(self):
        return list(self.name_var_scheme.values())

    def getNameVarSpecification(self):
        return list(self.name_var_specification.keys())

    def getProductNumber(self):
        return list(self.product_number.values())
